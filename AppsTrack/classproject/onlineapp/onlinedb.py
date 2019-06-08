import click
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()
from onlineapp.models import *
@click.group()
@click.option('-args')
def cli(args):
    pass

@cli.command()
def importdata():
    # conn = MySQLdb.connect('localhost', user='root', password='Fayaz#1233')
    # cursor = conn.cursor()
    # cursor.execute('use summer')
    # wb = load_workbook('students.xlsx')
    # ws = wb['Current']
    # for i in range(2, ws.max_row + 1):
    #     nm = str(ws.cell(row=i, column=1).value).lower()
    #     val = Student(name=ws.cell(row=i, column=1).value, email=str(ws.cell(row=i, column=3).value),
    #                   db_folder=str(ws.cell(row=i, column=4).value).lower(),
    #                   college_id=str(ws.cell(row=i, column=2).value))
    #
    #     val.save()

    # wb = load_workbook('students.xlsx')
    # ws = wb['Current']
    # for i in range(2, ws.max_row + 1):
    #     nm = str(ws.cell(row=i, column=1).value).lower()
    #     val = Student(name = ws.cell(row=i, column=1).value,email = str(ws.cell(row=i, column=3).value),db_folder = str(ws.cell(row=i, column=4).value).lower(),
    #            college_id = str(ws.cell(row=i, column=2).value))
    #
    #     val.save()

    wb = load_workbook('webscraping.xlsx')
    ws = wb['Sheet']
    for i in range(2,ws.max_row + 1):
        c=0
        name = str(ws.cell(row=i,column=2).value)
        ind = 0
        for j in range(len(name)):
            if name[j]=='_':
                c+=1
            if c==2:
                ind = j
                break
        name = name[ind+1:-5]
        name = name.lower()
        print(name)
        #print(ws.cell(row=2,column=3).value,ws.cell(row=2,column=4).value,ws.cell(row=2,column=5).value,ws.cell(row=2,column=6).value,ws.cell(row=2,column=7).value,name)
        val = MockTest(problem1=int(ws.cell(row=i,column=3).value),problem2=int(ws.cell(row=i,column=4).value),problem3=int(ws.cell(row=i,column=5).value),problem4=int(ws.cell(row=i,column=6).value),total=int(ws.cell(row=i,column=7).value),student_id=name)
        val.save()
        #cursor.execute('insert into mock(No ,Student,transform,from_custom_base26,get_pig_latin,top_charts,Total) values(%s,%s,%s,%s,%s,%s,%s)',val)

        #cursor.execute('insert into students values(%s,%s,%s,%s)',val)

    # conn.commit()
    pass

if __name__ == '__main__':
    cli()